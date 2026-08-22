"""真實持倉來源層：帳戶清單（永豐 Shioaji API＋任意個檔案帳戶）或純 ARK 模式。

各 skill 只透過 read_positions() 取得**全帳戶合併**的 {代號: (股數, 均價)}，
不知道來源是什麼。帳戶清單存於 ~/.ark-toolkit/config.json（僅非機密資料；
憑證在 .env，見 ark-setup）。回傳 None 代表「純 ARK 模式（不對帳）」——
與空 dict（持倉為零）語意不同，不可混用。

兩階段同步的交接快照（staging.json）也在這層：ark-collect 收集確認後寫入，
ark-sync 在有檔案帳戶時只讀快照——確認過的才是實際套用的。
"""
import csv
import datetime as dt
import json
import math
import os

CONFIG_PATH = os.path.expanduser("~/.ark-toolkit/config.json")
CREDENTIALS_FALLBACK = os.path.expanduser("~/.ark-toolkit/.env")
STAGING_PATH = os.path.expanduser("~/.ark-toolkit/staging.json")

ACCOUNT_TYPES = ("shioaji", "file")

# ARK 均價口徑（config.json 的 cost_basis）。缺欄位＝券商原值，舊設定不必遷移。
# 買進端沒有交易稅，「含稅」實際上就是手續費；股息是已領的現金股利（分錄的 ex_dividends）。
COST_BASIS_DEFAULT = {"include_dividends": False, "include_fees": False}
COST_BASIS_CHOICES = ({"include_dividends": False, "include_fees": False},
                      {"include_dividends": False, "include_fees": True},
                      {"include_dividends": True, "include_fees": False},
                      {"include_dividends": True, "include_fees": True})
# 「Σ分錄金額 ÷ 股數」與券商均價的容差：兩者各自四捨五入到分，差距必小於 0.01
LOT_CHECK_TOL = 0.01

# 常見券商匯出表頭 → 欄位自動對應（比對時忽略前後空白與大小寫）
CODE_HEADERS = ("股票代號", "證券代號", "商品代號", "股票代碼", "代號", "代碼", "code", "symbol")
QTY_HEADERS = ("庫存股數", "持有股數", "股數", "數量", "qty", "quantity", "shares")
PRICE_HEADERS = ("成交均價", "成本均價", "平均成本", "均價", "price", "avg_price")


class SetupRequired(RuntimeError):
    pass


class StagingRequired(RuntimeError):
    pass


def _validate_accounts(accounts):
    names = [a.get("name") for a in accounts]
    dupes = sorted({n for n in names if names.count(n) > 1})
    if dupes:
        raise RuntimeError(f"config.json 的帳戶名稱重複：{'、'.join(map(str, dupes))}")
    for a in accounts:
        if a.get("type") not in ACCOUNT_TYPES:
            raise RuntimeError(f"帳戶「{a.get('name')}」型別不明：{a.get('type')!r}"
                               f"（應為 {'/'.join(ACCOUNT_TYPES)}）")


def migrate_config(cfg):
    """v1（單一 source）→ v2（帳戶清單）。已是 v2 就驗證後原樣回傳。"""
    if "accounts" in cfg:
        _validate_accounts(cfg["accounts"])
        return cfg
    s = cfg.get("source")
    if s == "shioaji":
        accounts = [{"type": "shioaji", "name": "永豐"}]
    elif s == "csv":
        c = cfg.get("csv") or {}
        accounts = [{"type": "file", "name": "CSV",
                     "path": c["path"], "columns": c["columns"]}]
    elif s == "none":
        accounts = []
    else:
        raise RuntimeError(f"config.json 的 source 不明：{s!r}（應為 shioaji/csv/none）")
    out = {k: v for k, v in cfg.items() if k not in ("source", "csv")}
    out.update({"version": 2, "accounts": accounts})
    return out


def load_config(path=CONFIG_PATH):
    if not os.path.exists(path):
        raise SetupRequired("尚未設定真實持倉來源。請執行 ark-setup skill 完成初始設定。")
    with open(path, encoding="utf-8") as fh:
        cfg = json.load(fh)
    migrated = migrate_config(cfg)
    if migrated is not cfg:          # 舊格式：遷移後寫回，之後只有 v2 一種格式
        save_config(migrated, path)
    return migrated


def is_pure_ark(cfg):
    """純 ARK 模式（帳戶清單為空，不對帳）"""
    return not cfg.get("accounts")


def needs_staging(cfg):
    """有任何檔案帳戶就走兩階段（ark-collect 快照）；只有永豐可即時讀"""
    return any(a["type"] == "file" for a in cfg.get("accounts", []))


def save_config(config, path=CONFIG_PATH):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(config, fh, ensure_ascii=False, indent=2)
    return path


def auto_map_columns(headers):
    """從 CSV 表頭猜出代號／股數／均價欄位，認不出的為 None。"""
    def pick(candidates):
        normalized = {h.strip().lower(): h for h in headers}
        for c in candidates:
            if c.lower() in normalized:
                return normalized[c.lower()]
        return None
    return {"code": pick(CODE_HEADERS), "qty": pick(QTY_HEADERS), "price": pick(PRICE_HEADERS)}


def _to_number(s):
    return float(str(s).replace(",", "").strip())


def _read_rows(path):
    """依序以 UTF-8(BOM)、Big5(cp950) 嘗試解碼——台灣券商匯出常是 Big5。"""
    for enc in ("utf-8-sig", "cp950"):
        try:
            with open(path, encoding=enc, newline="") as fh:
                reader = csv.DictReader(fh)
                return reader.fieldnames or [], list(reader)
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"無法以 UTF-8 或 Big5(cp950) 解碼：{path}")


def _cell_str(v):
    """Excel 儲存格值 → 文字。整數值的浮點（代號 2330 常被存成 2330.0）去掉小數點。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _read_xlsx_rows(path):
    """讀 .xlsx 第一個工作表（首列為表頭），回傳與 _read_rows 相同的形狀。"""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        it = wb.worksheets[0].iter_rows(values_only=True)
        fieldnames = [_cell_str(v) for v in (next(it, None) or ()) if v is not None]
        rows = [{name: (_cell_str(r[i]) if i < len(r) else "")
                 for i, name in enumerate(fieldnames)} for r in it]
    finally:
        wb.close()
    return fieldnames, rows


def _dispatch_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return _read_rows(path)
    if ext == ".xlsx":
        return _read_xlsx_rows(path)
    raise RuntimeError(f"不支援的副檔名：{path}（僅 .csv／.xlsx）")


def sniff_headers(path):
    """讀出匯出檔（CSV／Excel）表頭，供 ark-setup 做欄位對應。"""
    return _dispatch_rows(path)[0]


def read_file_positions(path, columns):
    """讀券商匯出檔（CSV／Excel），回傳 {代號: (股數, 均價)}。

    解析失敗一律報錯，不靜默跳過——錯的持倉清單會讓 sync 做出錯的計畫。
    僅空代號與零股數列視為非持股資料而略過。
    """
    fieldnames, rows = _dispatch_rows(path)
    missing = [columns[k] for k in ("code", "qty", "price") if columns.get(k) not in fieldnames]
    if missing:
        raise RuntimeError(
            f"檔案缺少欄位 {missing}（實際表頭：{fieldnames}）。請重新執行 ark-setup 對應欄位。"
        )

    out = {}
    for i, row in enumerate(rows, start=2):
        code = (row.get(columns["code"]) or "").strip()
        if not code:
            continue
        try:
            qty_value = _to_number(row[columns["qty"]])
            price = _to_number(row[columns["price"]])
        except (ValueError, TypeError) as e:
            raise RuntimeError(f"第 {i} 行（{code}）數字解析失敗：{e}") from e
        if not math.isfinite(qty_value) or not qty_value.is_integer() or qty_value < 0:
            raise RuntimeError(f"第 {i} 行（{code}）股數必須是非負整數："
                               f"{row[columns['qty']]!r}")
        if not math.isfinite(price):
            raise RuntimeError(f"第 {i} 行（{code}）均價必須是有限數值："
                               f"{row[columns['price']]!r}")
        qty = int(qty_value)
        if qty == 0:
            continue
        if code in out:
            raise RuntimeError(f"第 {i} 行（{code}）重複出現，請先合併同一檔的持倉")
        out[code] = (qty, price)

    if not out:
        raise RuntimeError(f"沒有任何有效持股列：{path}")
    return out


def read_csv_positions(path, columns):
    """相容包裝：CSV 是 read_file_positions 支援格式之一。"""
    return read_file_positions(path, columns)


def merge_positions(by_account):
    """多帳戶合併：股數相加、均價加權平均（Σ股數×均價 ÷ Σ股數）。"""
    totals = {}
    for positions in by_account.values():
        for code, (qty, price) in positions.items():
            q0, cost0 = totals.get(code, (0, 0.0))
            totals[code] = (q0 + qty, cost0 + qty * price)
    return {code: (q, cost / q) for code, (q, cost) in totals.items() if q}


def read_account_positions(account, basis):
    """讀單一帳戶的即時持倉。錯誤一律指名帳戶——多帳戶時才知道是誰壞了。

    `basis` 刻意不給預設值：1.3.0 給了 `None` 預設，collect 的既有呼叫點就
    靜默退回券商原值，使用者選的口徑從沒進過快照。必填讓漏傳的呼叫點當場炸。
    """
    if account["type"] == "shioaji":
        return read_shioaji_positions(cost_basis=basis)
    path = os.path.expanduser(account["path"])
    if not os.path.exists(path):
        raise RuntimeError(f"帳戶「{account['name']}」找不到檔案：{path}。"
                           "請執行 ark-collect 重新選檔，或 ark-setup 變更路徑。")
    try:
        return read_file_positions(path, account["columns"])
    except RuntimeError as e:
        raise RuntimeError(f"帳戶「{account['name']}」讀取失敗：{e}") from e


def read_positions(config=None, config_path=CONFIG_PATH):
    """回傳全帳戶合併的真實持倉 {代號: (股數, 均價)}；純 ARK 模式回傳 None。

    任一帳戶讀不到就是整體失敗——絕不部分合併，少一個帳戶的「合併結果」
    會讓 sync 判定要刪掉那個帳戶的所有持股。
    """
    cfg = migrate_config(config) if config is not None else load_config(config_path)
    accounts = cfg["accounts"]
    if not accounts:
        return None
    basis = cost_basis(cfg)
    return merge_positions({a["name"]: read_account_positions(a, basis) for a in accounts})


def cost_basis(config):
    """config 的均價口徑；缺欄位或缺旗標一律補成券商原值（False）。"""
    return {**COST_BASIS_DEFAULT, **(config.get("cost_basis") or {})}


def describe_cost_basis(basis):
    return ("含息" if basis["include_dividends"] else "不含息") + "、" + \
           ("含手續費" if basis["include_fees"] else "不含手續費")


def has_shioaji(config):
    return any(a["type"] == "shioaji" for a in config.get("accounts", []))


def describe(config):
    """來源的簡短人話描述，供輸出訊息使用。

    口徑只對 Shioaji 帳戶有意義（檔案帳戶只有一欄均價，無從換算），
    所以只在有 Shioaji 時附上，免得檔案帳戶的使用者以為口徑有套用。
    """
    cfg = migrate_config(config)
    accounts = cfg["accounts"]
    if not accounts:
        return "純 ARK 模式（不對帳）"
    kind = {"shioaji": "Shioaji", "file": "檔案"}
    text = "＋".join(f"{a['name']}（{kind[a['type']]}）" for a in accounts)
    if has_shioaji(cfg):
        text += f"｜均價口徑：{describe_cost_basis(cost_basis(cfg))}"
    return text


# ---------------------------------------------------------------- 收集快照

def write_staging(by_account, path=STAGING_PATH, now=None):
    """把 ark-collect 確認過的各帳戶持倉寫成快照，含合併結果。

    任一帳戶 0 檔即拒寫——0 檔多半是讀取失敗，寫進快照會在 sync 變成
    「刪光該帳戶持股」的計畫。與 sync_is_safe 防同一類事，這裡守在更上游。
    """
    empty = sorted(name for name, pos in by_account.items() if not pos)
    if not by_account or empty:
        raise RuntimeError(f"帳戶「{'、'.join(empty) or '？'}」讀到 0 檔"
                           "（多半是讀取失敗），拒絕寫入快照")
    data = {
        "created_at": (now or dt.datetime.now()).replace(microsecond=0).isoformat(),
        "accounts": {name: {c: [q, p] for c, (q, p) in pos.items()}
                     for name, pos in by_account.items()},
        "merged": {c: [q, p] for c, (q, p) in merge_positions(by_account).items()},
    }
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)
    return path


def load_staging(path=STAGING_PATH, today=None):
    """讀回收集快照。不存在或非當日一律拒絕——重收很便宜，套用舊資料很貴。"""
    if not os.path.exists(path):
        raise StagingRequired("尚無收集快照。請先執行 ark-collect 收集並確認各帳戶持倉。")
    with open(path, encoding="utf-8") as fh:
        data = json.load(fh)
    created = dt.datetime.fromisoformat(data["created_at"])
    if created.date() != (today or dt.date.today()):
        raise StagingRequired(f"收集快照是 {created.date().isoformat()} 的"
                              "（僅當日有效），請重新執行 ark-collect。")
    def to_pos(d):
        return {c: (int(q), float(p)) for c, (q, p) in d.items()}
    return {"created_at": data["created_at"],
            "accounts": {name: to_pos(pos) for name, pos in data["accounts"].items()},
            "merged": to_pos(data["merged"])}


# ---------------------------------------------------------------- Shioaji

def load_credentials():
    """依序尋找憑證：工作目錄往上的 .env → ~/.ark-toolkit/.env → 既有環境變數。"""
    from dotenv import find_dotenv, load_dotenv

    found = find_dotenv(usecwd=True)
    if found:
        load_dotenv(found)
    if os.path.exists(CREDENTIALS_FALLBACK):
        load_dotenv(CREDENTIALS_FALLBACK)
    missing = [k for k in ("SHIOAJI_API_KEY", "SHIOAJI_SECRET_KEY") if not os.environ.get(k)]
    if missing:
        raise RuntimeError(
            f"缺少 {'、'.join(missing)}。請執行 ark-setup 設定，"
            f"或放進專案的 .env、{CREDENTIALS_FALLBACK}。"
        )


def lot_totals(details):
    """持倉分錄（list_position_detail）→ (Σ金額, Σ手續費, Σ已領股息)。

    實測（2026-08-22，永豐零股）：分錄的 `price` 是**該筆總金額**不是單價，
    `quantity` 對零股一律回 0——所以這裡只加總金額，股數要用 list_positions 的。
    """
    return (float(sum(d.price for d in details)),
            float(sum(d.fee for d in details)),
            float(sum(d.ex_dividends for d in details)))


def adjust_price(code, qty, raw_price, totals, basis):
    """依口徑換算均價：(Σ金額 ＋ 手續費? － 已領股息?) ÷ 股數，四捨五入到分。

    先驗算 Σ金額 ÷ 股數 ≈ 券商均價：整張持倉的分錄語意沒有樣本驗證過，
    對不上代表我們看錯了欄位，寧可中止 sync 也不把錯的數字寫進 ARK
    （與 check_parsed 驗總成本同一哲學——失敗不能偽裝成成功）。
    """
    if not basis["include_dividends"] and not basis["include_fees"]:
        return raw_price
    amount, fee, dividends = totals
    if qty <= 0 or abs(amount / qty - raw_price) >= LOT_CHECK_TOL:
        raise RuntimeError(f"{code} 的分錄金額 {amount:.0f} ÷ {qty} 股與券商均價 {raw_price} "
                           "對不上，無法換算口徑；請改回券商原值或回報此案例")
    cost = amount + (fee if basis["include_fees"] else 0.0) \
                  - (dividends if basis["include_dividends"] else 0.0)
    return round(cost / qty, 2)


def read_shioaji_positions(api=None, cost_basis=None):
    """回傳 {代號: (股數, 成本均價)}，均價依口徑換算（預設券商原值）。

    可傳入現成的 api session（ark-agent 一次要取多種資料）避免重複登入；
    不傳則維持原行為自行登入。券商原值時不查分錄——多一趟 API 沒有意義。
    只在自行登入時才 import shioaji：給了 api 的路徑（含測試的假 api）
    不該依賴套件是否安裝；`unit` 用文件明載的字面值 "Share"。
    """
    if api is None:
        import shioaji as sj

        load_credentials()
        api = sj.Shioaji(simulation=False)
        api.login(api_key=os.environ["SHIOAJI_API_KEY"],
                  secret_key=os.environ["SHIOAJI_SECRET_KEY"])
    basis = {**COST_BASIS_DEFAULT, **(cost_basis or {})}
    adjusting = basis["include_dividends"] or basis["include_fees"]
    out = {}
    for p in api.list_positions(api.stock_account, unit="Share"):
        qty, raw = int(p.quantity), float(p.price)
        if adjusting:
            totals = lot_totals(api.list_position_detail(api.stock_account, detail_id=p.id))
            out[p.code] = (qty, adjust_price(p.code, qty, raw, totals, basis))
        else:
            out[p.code] = (qty, raw)
    return out
